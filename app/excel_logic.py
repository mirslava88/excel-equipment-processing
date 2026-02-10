# Логика обработки Excel файлов
import pandas as pd
import tempfile
import os
import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime


CURRENT_YEAR = 2026
TECH_REFRESH_YEARS = 5
CRITICAL_AGE_YEARS = 9  # Возраст для критического устаревания


def _pluralize_years(n: int) -> str:
    """Склонение слова 'год/года/лет'."""
    if 11 <= n % 100 <= 19:
        return f"{n} лет"
    last_digit = n % 10
    if last_digit == 1:
        return f"{n} год"
    elif 2 <= last_digit <= 4:
        return f"{n} года"
    else:
        return f"{n} лет"


# Ключевые слова для автоопределения столбцов (в нижнем регистре)
SERIAL_KEYWORDS = [
    "серийный номер", "серийный", "серийник", "сер. номер", "сер.номер",
    "serial", "serial number", "serial_number", "serialnumber", "sn", "s/n", "с/н",
]
DATE_KEYWORDS = [
    "дата отражения проводки", "дата проводки", "дата отражения",
    "дата ввода", "дата внесения", "дата поступления", "дата",
    "posting date", "entry date", "date",
]


def get_engine(filename: str):
    """Определяет engine для pandas.read_excel по расширению файла."""
    ext = os.path.splitext(filename)[-1].lower()
    if ext == ".xlsb":
        return "pyxlsb"
    elif ext == ".xlsx":
        return "openpyxl"
    return None


def save_temp_file(upload_file) -> str:
    """Сохраняет загруженный файл во временную директорию."""
    ext = os.path.splitext(upload_file.filename)[-1]
    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        tmp.write(upload_file.file.read())
        return tmp.name


def _get_sheets_from_zip(filepath: str) -> list:
    """Низкоуровневое чтение листов из XLSX через ZIP и XML.
    Работает с любыми namespace, включая strict OOXML.
    """
    try:
        with zipfile.ZipFile(filepath, 'r') as z:
            if 'xl/workbook.xml' not in z.namelist():
                return []
            
            content = z.read('xl/workbook.xml')
            root = ET.fromstring(content)
            
            # Ищем все элементы, заканчивающиеся на 'sheet'
            all_elements = root.findall('.//')
            sheet_elements = [el for el in all_elements if el.tag.endswith('sheet') or el.tag.endswith('Sheet')]
            
            sheets = []
            for el in sheet_elements:
                name = el.get('name') or el.get('Name')
                if name:
                    sheets.append(name)
            
            return sheets
    except Exception:
        return []


def get_sheet_names(filepath: str, engine) -> list:
    """Возвращает список имён листов Excel-файла.
    Использует несколько методов для максимальной совместимости.
    """
    sheets = []
    
    # Метод 0: Низкоуровневое чтение ZIP (для strict OOXML и проблемных файлов)
    if engine == "openpyxl":
        sheets = _get_sheets_from_zip(filepath)
        if sheets:
            return sheets
    
    # Метод 1: openpyxl с разными параметрами (для xlsx)
    if engine == "openpyxl":
        try:
            import openpyxl
            # Попытка 1: read_only=False, keep_vba=False
            wb = openpyxl.load_workbook(filepath, read_only=False, keep_vba=False, data_only=False)
            sheets = [s for s in wb.sheetnames if s]  # Фильтруем пустые имена
            wb.close()
            if sheets:
                return sheets
        except Exception:
            pass
        
        try:
            # Попытка 2: read_only=True, keep_vba=True
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, keep_vba=True)
            sheets = [s for s in wb.sheetnames if s]
            wb.close()
            if sheets:
                return sheets
        except Exception:
            pass
    
    # Метод 2: pd.ExcelFile с указанным engine
    try:
        xls = pd.ExcelFile(filepath, engine=engine)
        sheets = [s for s in xls.sheet_names if s]
        xls.close()
        if sheets:
            return sheets
    except Exception:
        pass
    
    # Метод 3: pd.ExcelFile без указания engine (автодетект)
    try:
        xls = pd.ExcelFile(filepath)
        sheets = [s for s in xls.sheet_names if s]
        xls.close()
        if sheets:
            return sheets
    except Exception:
        pass
    
    # Метод 4: pd.read_excel с sheet_name=None (читает все листы)
    try:
        all_sheets = pd.read_excel(filepath, sheet_name=None, nrows=0, engine=engine)
        sheets = list(all_sheets.keys())
        if sheets:
            return sheets
    except Exception:
        pass
    
    raise Exception(f"Не удалось прочитать листы файла. Возможно, файл повреждён или имеет нестандартный формат.")


def _get_columns_from_zip(filepath: str, sheet_name: str) -> list:
    """Низкоуровневое чтение столбцов первой строки из XLSX через ZIP."""
    try:
        with zipfile.ZipFile(filepath, 'r') as z:
            # Находим соответствие имени листа и файла sheet?.xml
            workbook_content = z.read('xl/workbook.xml')
            wb_root = ET.fromstring(workbook_content)
            
            # Находим sheet элемент с нужным именем
            all_elements = wb_root.findall('.//')
            sheet_elements = [el for el in all_elements if el.tag.endswith('sheet')]
            
            rid = None
            for el in sheet_elements:
                name = el.get('name')
                if name == sheet_name:
                    # Находим relationship ID
                    for attr_name, attr_value in el.attrib.items():
                        if 'id' in attr_name.lower() and attr_value.startswith('rId'):
                            rid = attr_value
                            break
                    break
            
            if not rid:
                return []
            
            # Читаем relationships чтобы найти путь к sheet XML
            rels_content = z.read('xl/_rels/workbook.xml.rels')
            rels_root = ET.fromstring(rels_content)
            
            target_path = None
            for rel in rels_root.findall('.//*'):
                if rel.get('Id') == rid:
                    target_path = 'xl/' + rel.get('Target')
                    break
            
            if not target_path or target_path not in z.namelist():
                return []
            
            # Читаем sheet XML
            sheet_content = z.read(target_path)
            sheet_root = ET.fromstring(sheet_content)
            
            # Находим первую строку (row)
            all_elements = sheet_root.findall('.//')
            rows = [el for el in all_elements if el.tag.endswith('row')]
            
            if not rows:
                return []
            
            # Читаем sharedStrings если есть
            shared_strings = []
            if 'xl/sharedStrings.xml' in z.namelist():
                ss_content = z.read('xl/sharedStrings.xml')
                ss_root = ET.fromstring(ss_content)
                si_elements = [el for el in ss_root.findall('.//')if el.tag.endswith('si')]
                for si in si_elements:
                    t_elements = [el for el in si.findall('.//')if el.tag.endswith('t')]
                    if t_elements:
                        shared_strings.append(t_elements[0].text or '')
                    else:
                        shared_strings.append('')
            
            # Читаем ячейки первой строки
            first_row = rows[0]
            cells = [el for el in first_row.findall('.//')if el.tag.endswith('c')]
            
            columns = []
            for i, cell in enumerate(cells, 1):
                cell_type = cell.get('t')
                v_elements = [el for el in cell.findall('.//')if el.tag.endswith('v')]
                
                if v_elements and v_elements[0].text:
                    value = v_elements[0].text
                    # Если тип 's' - это индекс в sharedStrings
                    if cell_type == 's' and value.isdigit():
                        idx = int(value)
                        if idx < len(shared_strings):
                            columns.append(shared_strings[idx])
                        else:
                            columns.append(f"Column_{i}")
                    else:
                        columns.append(str(value))
                else:
                    columns.append(f"Column_{i}")
            
            return columns if columns else []
            
    except Exception:
        return []


def get_columns(filepath: str, engine, sheet_name: str) -> list:
    """Возвращает список столбцов указанного листа."""
    # Попытка 0: Низкоуровневое чтение из ZIP (для strict OOXML)
    if engine == "openpyxl":
        cols = _get_columns_from_zip(filepath, sheet_name)
        if cols:
            return cols
    
    # Попытка 1: pandas
    try:
        df = pd.read_excel(filepath, engine=engine, sheet_name=sheet_name, nrows=0)
        return [str(c) for c in df.columns.tolist()]
    except Exception:
        pass
    
    # Попытка 2: pandas без указания engine
    try:
        df = pd.read_excel(filepath, sheet_name=sheet_name, nrows=0)
        return [str(c) for c in df.columns.tolist()]
    except Exception:
        pass
    
    # Попытка 3: openpyxl напрямую (для проблемных файлов)
    if engine == "openpyxl":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            # Ищем лист по имени
            ws = None
            for sheet in wb.worksheets:
                if sheet.title == sheet_name:
                    ws = sheet
                    break
            
            if ws and ws.max_row > 0:
                # Читаем первую строку как заголовки
                cols = [str(cell.value) if cell.value else f"Column_{i}" 
                        for i, cell in enumerate(ws[1], 1)]
                wb.close()
                return cols
            wb.close()
        except Exception:
            pass
    
    raise Exception(f"Не удалось прочитать столбцы листа '{sheet_name}'")


def _match_column(columns: list, keywords: list) -> str:
    """Ищет первый столбец, название которого содержит одно из ключевых слов.
    Сначала пробует точное совпадение (без учёта регистра и пробелов),
    затем — частичное вхождение.
    """
    cols_lower = {str(c).strip().lower(): str(c) for c in columns}
    # Точное совпадение
    for kw in keywords:
        if kw in cols_lower:
            return cols_lower[kw]
    # Частичное вхождение (приоритет по порядку keywords)
    for kw in keywords:
        for col_low, col_orig in cols_lower.items():
            if kw in col_low:
                return col_orig
    return ""


def auto_detect_columns(columns: list) -> dict:
    """Автоопределяет столбцы серийных номеров и даты по ключевым словам.
    Возвращает dict: {"serial": "...", "date": "..."}.
    """
    return {
        "serial": _match_column(columns, SERIAL_KEYWORDS),
        "date": _match_column(columns, DATE_KEYWORDS),
    }


def _read_sheet_safe(filepath: str, engine, sheet_name: str) -> pd.DataFrame:
    """Безопасное чтение листа Excel с fallback для проблемных файлов."""
    # Попытка 0: calamine engine для strict OOXML (лучший вариант)
    if engine == "openpyxl":
        try:
            return pd.read_excel(filepath, engine="calamine", sheet_name=sheet_name)
        except (ValueError, KeyError, ImportError):
            pass
        
        # Если не нашли по имени, пробуем по индексу с calamine
        try:
            sheets = get_sheet_names(filepath, engine)
            if sheet_name in sheets:
                idx = sheets.index(sheet_name)
                return pd.read_excel(filepath, sheet_name=idx, engine="calamine")
        except (ValueError, KeyError, ImportError, Exception):
            pass
    
    # Попытка 1: стандартное чтение по имени
    try:
        return pd.read_excel(filepath, engine=engine, sheet_name=sheet_name)
    except (ValueError, KeyError):
        pass
    
    # Попытка 2: читаем все листы и ищем нужный
    try:
        all_sheets = pd.read_excel(filepath, sheet_name=None, engine=engine)
        if sheet_name in all_sheets:
            return all_sheets[sheet_name]
        # Ищем с учётом регистра
        for name, df in all_sheets.items():
            if name.lower() == sheet_name.lower():
                return df
    except Exception:
        pass
    
    # Попытка 3: читаем без engine
    try:
        return pd.read_excel(filepath, sheet_name=sheet_name)
    except (ValueError, KeyError):
        pass
    
    # Попытка 4: читаем все листы без engine
    try:
        all_sheets = pd.read_excel(filepath, sheet_name=None)
        if sheet_name in all_sheets:
            return all_sheets[sheet_name]
        for name, df in all_sheets.items():
            if name.lower() == sheet_name.lower():
                return df
    except Exception:
        pass
    
    # Попытка 5: используем индекс листа (получаем список листов и находим индекс)
    try:
        sheets = get_sheet_names(filepath, engine)
        if sheet_name in sheets:
            idx = sheets.index(sheet_name)
            return pd.read_excel(filepath, sheet_name=idx, engine=engine)
    except Exception:
        pass
    
    # Попытка 6: последний шанс - без engine по индексу
    try:
        sheets = get_sheet_names(filepath, engine)
        if sheet_name in sheets:
            idx = sheets.index(sheet_name)
            return pd.read_excel(filepath, sheet_name=idx)
    except Exception:
        pass
    
    raise Exception(f"Не удалось прочитать лист '{sheet_name}' из файла")


def process_excels(
    path1: str,
    path2: str,
    engine1,
    engine2,
    sheet1: str,
    sheet2: str,
    serial_col1: str,
    serial_col2: str,
    date_col1: str,
    date_col2: str,
    compare: bool = True,
    tech_refresh: bool = True,
) -> str:
    """
    Основная логика:
    1. Сравнивает серийные номера из двух листов (если compare=True).
    2. Добавляет столбец 'Передано на склад' (если compare=True).
    3. Если tech_refresh=True — сравнивает серийники с базой данных, 
       берет дату из базы и определяет устаревание (>5 лет).
    Возвращает путь к результирующему .xlsx файлу.
    """
    df1 = _read_sheet_safe(path1, engine1, sheet1)
    df2 = _read_sheet_safe(path2, engine2, sheet2)

    # Сверка серийных номеров (опционально)
    if compare:
        # Читаем лист "Возврат" из базы данных для сверки
        try:
            df_return = _read_sheet_safe(path2, engine2, "Возврат")
            
            # Приводим серийные номера к строковому типу и убираем пробелы
            df1[serial_col1] = df1[serial_col1].astype(str).str.strip().str.lower()
            
            # Проверяем, есть ли столбец serial_col2 на листе "Возврат"
            if serial_col2 in df_return.columns:
                df_return[serial_col2] = df_return[serial_col2].astype(str).str.strip().str.lower()
                serials_on_stock = set(df_return[serial_col2])
            else:
                # Если столбца нет, пробуем автоопределить
                detected = auto_detect_columns(df_return.columns.tolist())
                if detected["serial"]:
                    df_return[detected["serial"]] = df_return[detected["serial"]].astype(str).str.strip().str.lower()
                    serials_on_stock = set(df_return[detected["serial"]])
                else:
                    # Если не удалось определить, считаем что на складе ничего нет
                    serials_on_stock = set()
            
            # Векторизованное сравнение серийных номеров
            df1["Передано на склад"] = df1[serial_col1].isin(serials_on_stock).map(
                {True: "Да", False: "Нет"}
            )
        except Exception:
            # Если лист "Возврат" не найден или ошибка чтения
            df1["Передано на склад"] = "Нет (лист 'Возврат' не найден)"

    # Техрефреш оборудования (опционально)
    if tech_refresh and date_col2 and date_col2 in df2.columns:
        # Нормализуем серийные номера в обоих файлах
        df1_serials = df1[serial_col1].astype(str).str.strip().str.lower()
        df2_serials = df2[serial_col2].astype(str).str.strip().str.lower()
        
        # Создаем маппинг: серийный номер -> дата из базы данных
        serial_to_date = dict(zip(df2_serials, df2[date_col2]))
        
        # Инициализируем столбец
        df1["Оборудование устарело"] = "Не найдено в базе данных"
        
        # Для каждого серийника из файла обработки ищем дату в базе
        for idx, serial in enumerate(df1_serials):
            if serial in serial_to_date:
                date_val = serial_to_date[serial]
                
                # Извлекаем год из даты
                year = None
                if pd.notna(date_val):
                    if isinstance(date_val, (pd.Timestamp, datetime)):
                        year = date_val.year
                    else:
                        # Попытка парсинга строки
                        date_str = str(date_val).strip()
                        for fmt in ['%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y']:
                            try:
                                year = datetime.strptime(date_str, fmt).year
                                break
                            except:
                                continue
                
                if year:
                    age = CURRENT_YEAR - year
                    
                    if age <= TECH_REFRESH_YEARS:
                        df1.loc[idx, "Оборудование устарело"] = "Нет"
                    elif age <= CRITICAL_AGE_YEARS:
                        df1.loc[idx, "Оборудование устарело"] = f"Да, {_pluralize_years(age)}"
                    else:
                        df1.loc[idx, "Оборудование устарело"] = f"Критично, {_pluralize_years(age)}"

    out_path = os.path.join(tempfile.gettempdir(), "result.xlsx")
    df1.to_excel(out_path, index=False)
    return out_path
